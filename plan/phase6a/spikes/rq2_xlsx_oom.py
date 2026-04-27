"""RQ2 — openpyxl OOM on adversarial XLSX.

Generates a synthetic 100-sheet, 50K-row, 20-column XLSX (~20 MB target),
then iterates with ``load_workbook(read_only=True, data_only=True)`` and
applies the plan's per-sheet 50x30 cap (early break). Measures peak RSS
via ``resource.getrusage``.

PASS: peak RSS < 512 MB.
FAIL: lower row cap.

Run:
    /tmp/.spike6a-venv/bin/python plan/phase6a/spikes/rq2_xlsx_oom.py
"""

from __future__ import annotations

import os
import resource
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

# --- Tunables --------------------------------------------------------------
N_SHEETS = 100
N_ROWS = 50_000
N_COLS = 20
ROW_CAP = 50
COL_CAP = 30
# ---------------------------------------------------------------------------


def _peak_rss_mb() -> float:
    """Peak RSS in MB. macOS reports bytes, Linux reports KiB — handle both."""
    ru = resource.getrusage(resource.RUSAGE_SELF)
    raw = ru.ru_maxrss
    # Heuristic: > 1 GiB in KiB is unrealistic for this script; if the value
    # is suspiciously big assume bytes (Darwin).
    if sys.platform == "darwin":
        return raw / 1024 / 1024
    return raw / 1024  # Linux KiB → MiB


def make_xlsx(path: Path) -> int:
    """Build the adversarial workbook. Returns file size in bytes."""
    wb = Workbook(write_only=True)
    for s in range(N_SHEETS):
        ws = wb.create_sheet(f"sheet_{s:03d}")
        for r in range(N_ROWS):
            ws.append([f"r{r}c{c}" for c in range(N_COLS)])
        if s % 10 == 0:
            print(f"  generated sheet {s}/{N_SHEETS}", flush=True)
    wb.save(str(path))
    return path.stat().st_size


def extract_capped(path: Path) -> tuple[int, int, int]:
    """Mirror the planned extractor: read_only + cap 50x30/sheet.

    Returns ``(sheets, total_cells_read, total_chars)``.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets = 0
    cells = 0
    chars = 0
    try:
        for ws in wb.worksheets:
            sheets += 1
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= ROW_CAP:
                    break
                for c_idx, val in enumerate(row):
                    if c_idx >= COL_CAP:
                        break
                    cells += 1
                    if val is not None:
                        chars += len(str(val))
    finally:
        wb.close()
    return sheets, cells, chars


def main() -> int:
    with tempfile.TemporaryDirectory() as tmp:
        xlsx = Path(tmp) / "adversarial.xlsx"
        print("Generating adversarial XLSX (this is the slow part)...")
        t0 = time.monotonic()
        size = make_xlsx(xlsx)
        gen_dt = time.monotonic() - t0
        print(f"  size = {size / 1024 / 1024:.1f} MB ({gen_dt:.1f}s)")

        # Reset RSS measurement baseline by recording before the read.
        rss_before = _peak_rss_mb()
        print(f"  RSS before extraction: {rss_before:.1f} MB")

        t0 = time.monotonic()
        sheets, cells, chars = extract_capped(xlsx)
        ext_dt = time.monotonic() - t0
        rss_peak = _peak_rss_mb()

        print()
        print("=== RESULTS ===")
        print(f"file size:        {size / 1024 / 1024:.1f} MB")
        print(f"sheets:           {sheets}")
        print(f"cells read:       {cells}")
        print(f"chars extracted:  {chars}")
        print(f"extract time:     {ext_dt:.2f}s")
        print(f"peak RSS:         {rss_peak:.1f} MB")

        passed = rss_peak < 512
        verdict = "PASS" if passed else "FAIL"
        print(f"VERDICT: {verdict} (target < 512 MB)")
        return 0 if passed else 1


if __name__ == "__main__":
    sys.exit(main())
