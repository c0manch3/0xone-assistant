"""RQ2b — XLSX timing at the actual 20 MB Telegram cap.

The main RQ2 used 100 sheets × 50K rows × 20 cols, which produces a
308 MB file — *bigger* than the 20 MB Telegram pre-download limit.
The owner-relevant question: at the 20 MB limit, does openpyxl's
``read_only=True`` extraction with the 50×30 cap return in a
user-tolerable time?

We binary-search the row count to land just under 20 MB, then run the
same extractor.

Run:
    /tmp/.spike6a-venv/bin/python plan/phase6a/spikes/rq2b_xlsx_at_cap.py
"""

from __future__ import annotations

import resource
import sys
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROW_CAP = 50
COL_CAP = 30
TARGET_BYTES = 20 * 1024 * 1024


def _peak_rss_mb() -> float:
    raw = resource.getrusage(resource.RUSAGE_SELF).ru_maxrss
    return raw / 1024 / 1024 if sys.platform == "darwin" else raw / 1024


def make_xlsx(path: Path, n_sheets: int, n_rows: int, n_cols: int) -> int:
    wb = Workbook(write_only=True)
    for s in range(n_sheets):
        ws = wb.create_sheet(f"sheet_{s:03d}")
        for r in range(n_rows):
            ws.append([f"r{r}c{c}" for c in range(n_cols)])
    wb.save(str(path))
    return path.stat().st_size


def extract_capped(path: Path) -> tuple[int, int, int]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    sheets = cells = chars = 0
    try:
        for ws in wb.worksheets:
            sheets += 1
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= ROW_CAP:
                    break
                for c_idx, val in enumerate(row):
                    if c_idx >= COL_CAP:
                        break
                    cells += 1
                    if val is not None:
                        chars += len(str(val))
    finally:
        wb.close()
    return sheets, cells, chars


def main() -> int:
    # Empirical sweet spot for ~20 MB: 100 sheets × ~3000 rows × 20 cols
    # produces roughly 20 MB. Tune until under target.
    candidates = [
        (100, 3000, 20),
        (50, 6000, 20),
        (10, 30000, 20),
    ]
    with tempfile.TemporaryDirectory() as tmp:
        for n_sheets, n_rows, n_cols in candidates:
            path = Path(tmp) / f"sheets{n_sheets}_rows{n_rows}.xlsx"
            t0 = time.monotonic()
            size = make_xlsx(path, n_sheets, n_rows, n_cols)
            gen_dt = time.monotonic() - t0
            size_mb = size / 1024 / 1024
            print(f"\nshape: sheets={n_sheets} rows={n_rows} cols={n_cols}")
            print(f"  size = {size_mb:.1f} MB ({gen_dt:.1f}s gen)")
            if size > TARGET_BYTES * 1.5:
                print(f"  skip — over 30 MB target")
                continue

            t0 = time.monotonic()
            sheets, cells, chars = extract_capped(path)
            ext_dt = time.monotonic() - t0
            rss = _peak_rss_mb()
            print(f"  extract: {ext_dt:.2f}s  cells={cells}  chars={chars}  peak_rss={rss:.1f} MB")

    return 0


if __name__ == "__main__":
    sys.exit(main())
