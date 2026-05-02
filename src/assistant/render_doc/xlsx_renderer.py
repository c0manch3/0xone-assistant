"""Phase 9 §2.8 — XLSX renderer (openpyxl write_only over markdown
pipe-tables).

Pipeline:
  1. ``markdown_tables.parse(content_md)`` → list[Table].
  2. v1 constraint: exactly ONE table required (multi-sheet xlsx is
     §Явно НЕ #2).
  3. ``Workbook(write_only=True)`` (HIGH-4 — ~10× lower peak RSS than
     the default mode).
  4. Header row written via ``WriteOnlyCell`` with bold ``Font``.
  5. Body rows streamed via ``ws.append([WriteOnlyCell, ...])``.
  6. Per-sheet caps ``xlsx_max_rows`` (5000) + ``xlsx_max_cols`` (50).
  7. ``wb.save(final_path)``. Output cap ``xlsx_max_bytes`` checked
     post-render.

XLSX renderer does NOT depend on pandoc — openpyxl ships as pure-Python
wheel (already in pyproject.toml from phase 6a). HIGH-5 partial
force-disable: xlsx works even when pandoc is missing.
"""

from __future__ import annotations

import contextlib
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from assistant.config import RenderDocSettings
from assistant.logger import get_logger
from assistant.render_doc.markdown_tables import (
    MarkdownTableError,
)
from assistant.render_doc.markdown_tables import (
    parse as parse_tables,
)
from assistant.render_doc.pdf_renderer import PDFRenderError as _PDFRenderError

log = get_logger("render_doc.xlsx_renderer")

XLSXRenderError = _PDFRenderError


async def render_xlsx(
    content_md: str,
    *,
    final_path: Path,
    settings: RenderDocSettings,
) -> int:
    """Render ``content_md`` (single pipe-table) to XLSX at
    ``final_path``. Returns bytes written. Raises
    :class:`XLSXRenderError` on any failure."""
    import asyncio

    try:
        tables = parse_tables(content_md)
    except MarkdownTableError as exc:
        raise XLSXRenderError(
            "render_failed_input_syntax",
            exc.code,
            message="markdown table parse",
        ) from exc

    if len(tables) == 0:
        raise XLSXRenderError(
            "render_failed_input_syntax",
            "markdown-no-tables",
            message="content_md contains no pipe-table",
        )
    if len(tables) > 1:
        raise XLSXRenderError(
            "render_failed_input_syntax",
            "markdown-multi-table",
            message=f"v1 supports a single table; got {len(tables)}",
        )

    table = tables[0]
    n_cols = len(table.header)
    if n_cols > settings.xlsx_max_cols:
        raise XLSXRenderError(
            "render_failed_input_syntax",
            "openpyxl-too-many-cols",
            message=f"{n_cols} > {settings.xlsx_max_cols}",
        )
    if len(table.rows) > settings.xlsx_max_rows:
        raise XLSXRenderError(
            "render_failed_input_syntax",
            "openpyxl-too-many-rows",
            message=f"{len(table.rows)} > {settings.xlsx_max_rows}",
        )

    def _save_sync() -> int:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("Sheet1")
        # write_only mode returns a WriteOnlyWorksheet but openpyxl's
        # ``create_sheet`` is typed as the broader Worksheet union.
        # WriteOnlyCell + ws.append(...) are the only operations we
        # need; both are valid on the union.
        bold_font = Font(bold=True)
        from openpyxl.cell import WriteOnlyCell

        header_cells = []
        for h in table.header:
            cell = WriteOnlyCell(ws, value=h)
            cell.font = bold_font
            header_cells.append(cell)
        ws.append(header_cells)
        for row in table.rows:
            ws.append(row)
        wb.save(str(final_path))
        return final_path.stat().st_size

    try:
        bytes_out = await asyncio.to_thread(_save_sync)
    except Exception as exc:
        raise XLSXRenderError(
            "render_failed_internal",
            "openpyxl-error",
            message=f"{type(exc).__name__}: {exc!s}"[:256],
        ) from exc

    if bytes_out > settings.xlsx_max_bytes:
        with contextlib.suppress(OSError):
            final_path.unlink(missing_ok=True)
        raise XLSXRenderError(
            "render_failed_output_cap",
            "xlsx-too-large",
            message=f"{bytes_out} > {settings.xlsx_max_bytes}",
        )
    return bytes_out
