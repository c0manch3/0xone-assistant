"""Phase 6a — XLSX extractor unit tests.

Covers:
- happy path: 3-sheet workbook with sheet separators;
- ROW_CAP / COL_CAP enforcement (RQ2 / Q13 retune);
- POST_EXTRACT_CHAR_CAP clean-boundary (devil H1);
- BadZipFile → ExtractionError;
- ``data_only=True`` cached-formula caveat (devil L3).
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from assistant.files.extract import (
    POST_EXTRACT_CHAR_CAP,
    XLSX_COL_CAP,
    XLSX_ROW_CAP,
    ExtractionError,
    extract_xlsx,
)


def _make_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    """Build an XLSX with the given sheet→rows-of-cells map."""
    wb = Workbook()
    # Drop the default sheet we don't want.
    default = wb.active
    if default is not None:
        wb.remove(default)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    wb.save(str(path))
    wb.close()


def test_extract_xlsx_happy_path_three_sheets(tmp_path: Path) -> None:
    p = tmp_path / "three.xlsx"
    _make_workbook(
        p,
        {
            "Sheet1": [["a", 1], ["b", 2]],
            "Sheet2": [["x", "y"]],
            "Sheet3": [["only"]],
        },
    )

    text, n = extract_xlsx(p)
    assert n == len(text)
    assert "=== Sheet: Sheet1 ===" in text
    assert "=== Sheet: Sheet2 ===" in text
    assert "=== Sheet: Sheet3 ===" in text
    assert "a\t1" in text
    assert "x\ty" in text
    assert "only" in text


def test_extract_xlsx_row_cap_enforced(tmp_path: Path) -> None:
    """Sheet with 50 rows extracts exactly XLSX_ROW_CAP rows."""
    p = tmp_path / "many_rows.xlsx"
    rows = [[f"r{i}c0", f"r{i}c1"] for i in range(50)]
    _make_workbook(p, {"Big": rows})

    text, _ = extract_xlsx(p)
    # Header line + XLSX_ROW_CAP data rows.
    data_lines = [
        line for line in text.split("\n") if line.startswith("r") and "\t" in line
    ]
    assert len(data_lines) == XLSX_ROW_CAP


def test_extract_xlsx_col_cap_enforced(tmp_path: Path) -> None:
    """Row with > XLSX_COL_CAP cells truncates to XLSX_COL_CAP."""
    p = tmp_path / "wide.xlsx"
    wide_row = [f"c{i}" for i in range(50)]
    _make_workbook(p, {"Wide": [wide_row]})

    text, _ = extract_xlsx(p)
    # First data row is below the header.
    lines = text.split("\n")
    data_line = next(line for line in lines if line.startswith("c0\t"))
    cells = data_line.split("\t")
    assert len(cells) == XLSX_COL_CAP


def test_extract_xlsx_none_cell_renders_empty(tmp_path: Path) -> None:
    """``None`` cells render as empty string, not the literal ``"None"``."""
    p = tmp_path / "none.xlsx"
    _make_workbook(p, {"S": [["a", None, "c"]]})

    text, _ = extract_xlsx(p)
    assert "a\t\tc" in text
    assert "None" not in text


def test_extract_xlsx_total_cap_skips_remaining_sheets(tmp_path: Path) -> None:
    """Devil H1: when the running total would exceed POST_EXTRACT_CHAR_CAP,
    skip the remaining sheets at a clean boundary and surface a marker.
    """
    # Build a per-sheet payload of ~70 KB so three sheets fit (3 *
    # 70 KB = 210 KB > 200 KB cap) — the first two sheets land, the
    # third triggers the skip.
    fat_cell = "x" * 110  # 110 chars/cell × 30 cols × 20 rows ≈ 66 KB/sheet
    sheet_rows = [[fat_cell] * XLSX_COL_CAP for _ in range(XLSX_ROW_CAP)]
    p = tmp_path / "fat.xlsx"
    _make_workbook(
        p,
        {
            "S1": sheet_rows,
            "S2": sheet_rows,
            "S3": sheet_rows,
            "S4": sheet_rows,
        },
    )

    text, _ = extract_xlsx(p)
    assert "S1" in text
    assert "S2" in text
    # The cap message tells the model how many sheets were dropped.
    assert "truncated" in text
    assert "more sheet" in text
    # S3 / S4 should be reported as skipped.
    assert "S4" not in text
    # The total never exceeds POST_EXTRACT_CHAR_CAP + the truncate marker
    # (the marker is < 200 chars).
    assert len(text) <= POST_EXTRACT_CHAR_CAP + 500


def test_extract_xlsx_total_cap_first_sheet_oversize_handled(tmp_path: Path) -> None:
    """Edge case (defensive assert in handler): a single sheet's chunk
    exceeds the cap on its own. The extractor's clean-boundary
    truncation cannot help — the chunk is too big from the start. The
    handler then runs a final substring truncation; the extractor
    just emits the cap-skip marker and an empty body.

    This test pins the extractor-side behaviour: the marker carries
    the count of sheets skipped (here all 4), and no body content
    appears.
    """
    fat_cell = "x" * 1500  # 1500 × 30 × 20 ≈ 900 KB / sheet — far above cap
    sheet_rows = [[fat_cell] * XLSX_COL_CAP for _ in range(XLSX_ROW_CAP)]
    p = tmp_path / "huge.xlsx"
    _make_workbook(
        p,
        {
            "S1": sheet_rows,
            "S2": sheet_rows,
            "S3": sheet_rows,
            "S4": sheet_rows,
        },
    )

    text, _ = extract_xlsx(p)
    # No sheet header survives; the marker reports 4 sheets skipped.
    assert "truncated" in text
    assert "4 more sheet" in text
    # No actual cell content.
    assert "x" * 200 not in text


def test_extract_xlsx_corrupt_raises(tmp_path: Path) -> None:
    """A non-zip file → BadZipFile → ExtractionError("encrypted or…")."""
    p = tmp_path / "not_really.xlsx"
    p.write_bytes(b"\x00\x01 not a zip")

    with pytest.raises(ExtractionError) as excinfo:
        extract_xlsx(p)
    # Handler keys on "encrypted" substring for the Russian reply.
    assert "encrypted" in str(excinfo.value).lower() or "corrupt" in str(
        excinfo.value
    ).lower()


def test_extract_xlsx_missing_file_raises(tmp_path: Path) -> None:
    p = tmp_path / "ghost.xlsx"
    with pytest.raises(ExtractionError):
        extract_xlsx(p)


def test_extract_xlsx_data_only_caveat_no_formulas_cached(tmp_path: Path) -> None:
    """Devil L3: openpyxl-generated XLSX without an Excel save-pass shows
    None (rendered as ``""``) for formula cells. Documenting the
    behaviour as a contract — owner-uploaded real Excel files have
    cached values and render correctly.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = 5
    ws["A2"] = 10
    ws["A3"] = "=SUM(A1:A2)"  # formula, no cached value
    p = tmp_path / "formulas.xlsx"
    wb.save(str(p))
    wb.close()

    text, _ = extract_xlsx(p)
    # First two rows render their numeric values.
    assert "5" in text
    assert "10" in text
    # Formula cell renders as "" (data_only=True returns None).
    # Confirm the literal "=SUM" is NOT in output.
    assert "=SUM" not in text


def test_extract_xlsx_wb_close_called_in_finally(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """``wb.close()`` is invoked even if the iter_rows loop raises.

    Mandatory contract per RQ2 — openpyxl read-only workbooks leak fd
    if not explicitly closed.
    """
    p = tmp_path / "ok.xlsx"
    _make_workbook(p, {"S": [["a", "b"]]})

    closed = {"count": 0}
    from openpyxl import load_workbook as real_load_workbook

    def patched_load(*args, **kwargs):  # type: ignore[no-untyped-def]
        wb = real_load_workbook(*args, **kwargs)
        original_close = wb.close

        def tracking_close() -> None:
            closed["count"] += 1
            original_close()

        wb.close = tracking_close  # type: ignore[method-assign]

        # Replace iter_rows on the active worksheet to raise.
        ws = wb["S"]

        def explode(*a: object, **kw: object):  # type: ignore[no-untyped-def]
            raise RuntimeError("boom from iter_rows")

        ws.iter_rows = explode  # type: ignore[method-assign]
        return wb

    # ``extract_xlsx`` imports ``load_workbook`` inside the function
    # body; patching at the openpyxl module level intercepts the
    # lookup.
    import openpyxl

    monkeypatch.setattr(openpyxl, "load_workbook", patched_load)

    with pytest.raises(ExtractionError):
        extract_xlsx(p)
    assert closed["count"] == 1
