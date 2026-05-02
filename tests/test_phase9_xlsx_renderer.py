"""Phase 9 §2.8 — XLSX renderer (write_only mode + bold header) tests.

XLSX renderer doesn't depend on pandoc — pure-Python openpyxl. All
tests run unconditionally on every host.
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import load_workbook

from assistant.config import RenderDocSettings
from assistant.render_doc.pdf_renderer import PDFRenderError as RenderError
from assistant.render_doc.xlsx_renderer import render_xlsx


@pytest.mark.asyncio
async def test_xlsx_single_table_happy_path(tmp_path: Path) -> None:
    md = (
        "| Имя | Возраст |\n"
        "|-----|---------|\n"
        "| Виталий | 35 |\n"
        "| Other  | 22 |\n"
    )
    final_path = tmp_path / "out.xlsx"
    bytes_out = await render_xlsx(
        md, final_path=final_path, settings=RenderDocSettings()
    )
    assert bytes_out > 0
    assert final_path.exists()
    # Re-open and verify structure + Cyrillic header preserved.
    wb = load_workbook(final_path, read_only=True)
    ws = wb.active
    assert ws is not None
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Имя", "Возраст")
    assert rows[1] == ("Виталий", "35")
    assert rows[2] == ("Other", "22")


@pytest.mark.asyncio
async def test_xlsx_no_tables_returns_input_syntax(tmp_path: Path) -> None:
    final_path = tmp_path / "out.xlsx"
    with pytest.raises(RenderError) as ei:
        await render_xlsx(
            "Just prose, no table.",
            final_path=final_path,
            settings=RenderDocSettings(),
        )
    assert ei.value.reason == "render_failed_input_syntax"
    assert ei.value.error_code == "markdown-no-tables"


@pytest.mark.asyncio
async def test_xlsx_multi_table_rejected(tmp_path: Path) -> None:
    md = (
        "| a | b |\n|---|---|\n| 1 | 2 |\n\n"
        "| c | d |\n|---|---|\n| 3 | 4 |\n"
    )
    final_path = tmp_path / "out.xlsx"
    with pytest.raises(RenderError) as ei:
        await render_xlsx(
            md, final_path=final_path, settings=RenderDocSettings()
        )
    assert ei.value.error_code == "markdown-multi-table"


@pytest.mark.asyncio
async def test_xlsx_too_many_rows_rejected(tmp_path: Path) -> None:
    """v1 cap = 5000 rows; renderer rejects 5001-row table BEFORE
    write."""
    header = "| col |\n|-----|\n"
    body = "| x |\n" * 5001
    final_path = tmp_path / "out.xlsx"
    with pytest.raises(RenderError) as ei:
        await render_xlsx(
            header + body,
            final_path=final_path,
            settings=RenderDocSettings(),
        )
    assert ei.value.error_code == "openpyxl-too-many-rows"


@pytest.mark.asyncio
async def test_xlsx_too_many_cols_rejected(tmp_path: Path) -> None:
    """Default ``xlsx_max_cols=50`` — 51-col table rejected."""
    header_cells = " | ".join(f"c{i}" for i in range(51))
    sep_cells = "|".join("-" * 3 for _ in range(51))
    body_cells = " | ".join(str(i) for i in range(51))
    md = (
        f"| {header_cells} |\n"
        f"|{sep_cells}|\n"
        f"| {body_cells} |\n"
    )
    final_path = tmp_path / "out.xlsx"
    with pytest.raises(RenderError) as ei:
        await render_xlsx(
            md, final_path=final_path, settings=RenderDocSettings()
        )
    assert ei.value.error_code == "openpyxl-too-many-cols"


@pytest.mark.asyncio
async def test_xlsx_header_row_bold(tmp_path: Path) -> None:
    """HIGH-4 / AC#3: header row uses ``Font(bold=True)`` style."""
    md = "| col A | col B |\n|---|---|\n| 1 | 2 |\n"
    final_path = tmp_path / "out.xlsx"
    await render_xlsx(
        md, final_path=final_path, settings=RenderDocSettings()
    )
    # write_only mode + WriteOnlyCell font is preserved post-load.
    wb = load_workbook(final_path)
    ws = wb.active
    assert ws is not None
    header_a = ws.cell(row=1, column=1)
    assert header_a.font.bold is True
    body_a = ws.cell(row=2, column=1)
    # Body cells should NOT inherit the bold style.
    assert body_a.font.bold in (False, None)
