"""Phase 7 / commit 18j — extract_doc defusedxml + zip-bomb coverage.

The CLI at ``tools/extract_doc/main.py`` documents a dual-layer defence
for ZIP-backed Office formats (see its module docstring §2.9):

1. ``_guard_zip_bomb()`` rejects archives whose declared uncompressed
   total exceeds ``_ZIP_UNCOMPRESSED_CAP`` (64 MB) **before** any
   parser opens a part.
2. ``defusedxml.defuse_stdlib()`` is called at module load time so
   entity-expansion (billion-laughs) and external-entity (XXE)
   attacks inside DOCX / XLSX parts are intercepted.

This test file stress-tests both layers end-to-end via the
subprocess boundary (mirroring the Bash-allowlist runtime invocation)
and includes control cases to catch over-broad rejection.

PRODUCTION GAP (flagged, not fixed — phase-7 follow-up):
python-docx parses ``word/document.xml`` via lxml's ``XMLParser(
resolve_entities=False)``, not via stdlib ``xml.*``, so
``defuse_stdlib()`` in the CLI does NOT intercept hostile DOCX XML.
The lxml ``resolve_entities=False`` default keeps the CLI SAFE
(no leak, no expansion) but the CLI silently returns exit 0 with
empty text instead of rejecting. The hard safety invariants are
tested unconditionally; the explicit-rejection expectation is
marked ``xfail(strict=True)`` so it surfaces the day an explicit
defusedxml/lxml DOCTYPE-forbidding parser is wired in.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import pytest
from docx import Document
from openpyxl import Workbook

_CLI = Path(__file__).resolve().parents[1] / "tools" / "extract_doc" / "main.py"

# Must match ``_ZIP_UNCOMPRESSED_CAP`` in ``tools/extract_doc/main.py``.
# 70 x 1 MB = 70 MB declared > 64 MB cap.
_BOMB_ENTRY_COUNT = 70
_BOMB_ENTRY_SIZE = 1024 * 1024


def _run(*args: str) -> subprocess.CompletedProcess[str]:
    """Invoke the CLI in a fresh subprocess, matching Bash-allowlist usage."""
    return subprocess.run(
        [sys.executable, str(_CLI), *args],
        env=dict(os.environ),
        capture_output=True,
        text=True,
        check=False,
        timeout=30,
    )


# --- fixture builders -------------------------------------------------------


def _make_legit_docx(path: Path) -> None:
    """Write a minimal legit DOCX (control case)."""
    doc = Document()
    doc.add_paragraph("Control: this document is legitimate.")
    doc.save(str(path))


def _make_zip_bomb(path: Path) -> None:
    """Write a ZIP with ``_BOMB_ENTRY_COUNT`` x 1 MB zero-filled entries.

    DEFLATE on zero bytes compresses near-perfectly, so the file on
    disk is ~70 KB while the declared uncompressed total is ~70 MB —
    comfortably above the 64 MB cap. This models the real attack
    shape: tiny payload, huge memory balloon on naive extract.
    """
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        for i in range(_BOMB_ENTRY_COUNT):
            zf.writestr(f"ballast{i}.bin", b"\x00" * _BOMB_ENTRY_SIZE)


def _rebuild_docx_with_document_xml(source_docx: Path, dest: Path, document_xml: bytes) -> None:
    """Clone ``source_docx`` into ``dest`` with ``word/document.xml`` replaced.

    Preserves every other zip entry (content-types, rels, styles, ...)
    so python-docx's OPC parser walks the package correctly before
    reaching the hostile part. Without this, the parser fails on
    missing relationships instead of on the XML payload — which would
    test the wrong layer.
    """
    with (
        zipfile.ZipFile(source_docx) as zin,
        zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zout,
    ):
        for item in zin.infolist():
            data = document_xml if item.filename == "word/document.xml" else zin.read(item.filename)
            zout.writestr(item, data)


# XML payloads for entity-expansion / XXE attacks. Kept at module scope
# so they're easy to inspect and don't silently mutate between tests.

_BILLION_LAUGHS_DOCUMENT_XML = b"""<?xml version="1.0"?>
<!DOCTYPE lolz [
  <!ENTITY lol "lol">
  <!ENTITY lol2 "&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;&lol;">
  <!ENTITY lol3 "&lol2;&lol2;&lol2;&lol2;&lol2;&lol2;&lol2;&lol2;&lol2;&lol2;">
  <!ENTITY lol4 "&lol3;&lol3;&lol3;&lol3;&lol3;&lol3;&lol3;&lol3;&lol3;&lol3;">
]>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body><w:p><w:r><w:t>&lol4;</w:t></w:r></w:p></w:body>
</w:document>
"""


def _xxe_document_xml(target_file: Path) -> bytes:
    """Build a DOCX document.xml that tries to exfiltrate ``target_file``.

    A hardened parser MUST refuse the DOCTYPE outright (DTDForbidden /
    EntitiesForbidden). A merely silent-ignore strategy (e.g. lxml's
    ``resolve_entities=False``) prevents the secret from leaking but
    does NOT surface the attack — see the surprise note on
    ``test_xxe_docx_rejected_by_defusedxml``.
    """
    return (
        f'<?xml version="1.0"?>\n'
        f"<!DOCTYPE foo [\n"
        f'  <!ENTITY xxe SYSTEM "file://{target_file}">\n'
        f"]>\n"
        f'<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">\n'
        f"  <w:body><w:p><w:r><w:t>&xxe;</w:t></w:r></w:p></w:body>\n"
        f"</w:document>\n"
    ).encode()


# --- zip-bomb rejection -----------------------------------------------------


def test_zip_bomb_xlsx_rejected_before_parse(tmp_path: Path) -> None:
    """A .xlsx whose declared uncompressed total blows past the 64 MB cap
    must trip ``_guard_zip_bomb`` with exit 3 before openpyxl opens it.
    """
    bomb = tmp_path / "bomb.xlsx"
    _make_zip_bomb(bomb)

    result = _run(str(bomb))

    assert result.returncode == 3, (
        f"expected validation exit (3), got {result.returncode}\n"
        f"stdout={result.stdout!r}\nstderr={result.stderr!r}"
    )
    payload = json.loads(result.stderr)
    assert payload["ok"] is False
    assert "zip-bomb guard" in payload["error"]
    assert payload["declared"] >= 64 * 1024 * 1024
    assert payload["cap"] == 64 * 1024 * 1024


def test_zip_bomb_docx_rejected_before_parse(tmp_path: Path) -> None:
    """Same attack model as the XLSX case, but the extension routes
    through ``_extract_docx`` instead of ``_extract_xlsx``. Both paths
    must trip ``_guard_zip_bomb`` with identical semantics.
    """
    bomb = tmp_path / "bomb.docx"
    _make_zip_bomb(bomb)

    result = _run(str(bomb))

    assert result.returncode == 3, (
        f"expected validation exit (3), got {result.returncode}\n"
        f"stdout={result.stdout!r}\nstderr={result.stderr!r}"
    )
    payload = json.loads(result.stderr)
    assert payload["ok"] is False
    assert "zip-bomb guard" in payload["error"]
    assert payload["declared"] >= 64 * 1024 * 1024
    assert payload["cap"] == 64 * 1024 * 1024


# --- defusedxml rejection ---------------------------------------------------


def test_xxe_docx_does_not_leak_secret(tmp_path: Path) -> None:
    """HARD guarantee: an XXE DOCX must NEVER leak the external file
    contents into stdout/stderr.

    The hostile ``document.xml`` declares an external entity pointing
    at a local secret and references it inside ``<w:t>``. Regardless
    of whether the CLI rejects the file or silently emits an empty
    transcript, the secret MUST NOT appear anywhere in the CLI
    output — this is the ONLY non-negotiable security property
    (leakage would be a CVE-class regression).

    The separate rejection-semantics expectation is asserted in
    ``test_xxe_docx_rejected_by_defusedxml`` below (currently xfail;
    see the surprise note there).
    """
    legit = tmp_path / "ok.docx"
    _make_legit_docx(legit)

    secret = tmp_path / "secret.txt"
    secret_value = "TOP_SECRET_EXTRACT_DOC_CANARY_XYZZY"
    secret.write_text(secret_value, encoding="utf-8")

    evil = tmp_path / "xxe.docx"
    _rebuild_docx_with_document_xml(legit, evil, _xxe_document_xml(secret))

    result = _run(str(evil))

    assert secret_value not in result.stdout, (
        f"XXE leaked secret into stdout — CRITICAL regression; stdout={result.stdout!r}"
    )
    assert secret_value not in result.stderr, (
        f"XXE leaked secret into stderr — CRITICAL regression; stderr={result.stderr!r}"
    )


@pytest.mark.xfail(
    strict=True,
    reason=(
        "Production gap: python-docx parses document.xml with lxml's "
        "XMLParser(resolve_entities=False), not via stdlib xml, so "
        "defuse_stdlib() in tools/extract_doc/main.py does not intercept. "
        "Entities are silently unresolved (no leak, no expansion) but the "
        "CLI returns exit 0 with empty text — there is no explicit DOCTYPE/"
        "entity rejection. Flagged for phase-7 fix; do not fix in this test "
        "commit. See module docstring."
    ),
)
def test_xxe_docx_rejected_by_defusedxml(tmp_path: Path) -> None:
    """An XXE DOCX SHOULD be refused by the XML parser with a non-zero
    exit so callers can distinguish hostile input from legitimate
    empty documents. Currently xfail — production silently emits
    empty text (secret is NOT leaked; see the paired hard-guarantee
    test). Remove the xfail once explicit rejection lands.
    """
    legit = tmp_path / "ok.docx"
    _make_legit_docx(legit)

    secret = tmp_path / "secret.txt"
    secret.write_text("CANARY", encoding="utf-8")

    evil = tmp_path / "xxe.docx"
    _rebuild_docx_with_document_xml(legit, evil, _xxe_document_xml(secret))

    result = _run(str(evil))

    assert result.returncode != 0, (
        f"expected non-zero exit on XXE, got 0\nstdout={result.stdout!r}\nstderr={result.stderr!r}"
    )


def test_billion_laughs_docx_does_not_expand(tmp_path: Path) -> None:
    """HARD guarantee: a billion-laughs DOCX must NEVER have its
    entity expansion materialise into the CLI output.

    A successful attack would emit the word "lol" thousands of times;
    a safe parser (rejecting OR silently-ignoring) keeps the output
    clean. Either outcome is acceptable for THIS safety invariant;
    the rejection-semantics expectation lives in the paired xfail
    test below.
    """
    legit = tmp_path / "ok.docx"
    _make_legit_docx(legit)

    evil = tmp_path / "billion-laughs.docx"
    _rebuild_docx_with_document_xml(legit, evil, _BILLION_LAUGHS_DOCUMENT_XML)

    result = _run(str(evil))

    # Unexpanded output can legitimately contain a handful of "lol"
    # tokens from the entity *definition* echoed back in error text;
    # a runaway expansion produces thousands. 10 is a comfortable
    # ceiling that keeps the test robust while still catching the
    # attack.
    lol_count = result.stdout.count("lol") + result.stderr.count("lol")
    assert lol_count < 10, (
        "billion-laughs entities appear to have expanded — "
        "CRITICAL regression; "
        f"lol_count={lol_count}; "
        f"first 200 chars stdout={result.stdout[:200]!r}; "
        f"first 200 chars stderr={result.stderr[:200]!r}"
    )


@pytest.mark.xfail(
    strict=True,
    reason=(
        "Production gap: same root cause as test_xxe_docx_rejected_by_defusedxml. "
        "python-docx's lxml parser uses resolve_entities=False so billion-laughs "
        "entities stay unresolved (no memory blow-up) but the CLI returns exit 0 "
        "with empty text — no explicit DOCTYPE/entity rejection. Flagged for "
        "phase-7 fix."
    ),
)
def test_billion_laughs_docx_rejected_by_defusedxml(tmp_path: Path) -> None:
    """A billion-laughs DOCX SHOULD be refused by the XML parser with a
    non-zero exit. Currently xfail — see the paired hard-guarantee
    test for the non-negotiable safety invariant that DOES hold.
    """
    legit = tmp_path / "ok.docx"
    _make_legit_docx(legit)

    evil = tmp_path / "billion-laughs.docx"
    _rebuild_docx_with_document_xml(legit, evil, _BILLION_LAUGHS_DOCUMENT_XML)

    result = _run(str(evil))

    assert result.returncode != 0, (
        f"expected non-zero exit on billion-laughs, got 0\n"
        f"stdout={result.stdout!r}\nstderr={result.stderr!r}"
    )


# --- control case -----------------------------------------------------------


def test_legitimate_docx_still_passes(tmp_path: Path) -> None:
    """Control: a plain legit DOCX must still extract successfully.

    Guards against over-broad rejection (e.g. a hypothetical fix that
    rejected every DOCX containing a DOCTYPE would also false-positive
    on legitimate documents that happen to declare one).
    """
    src = tmp_path / "legit.docx"
    _make_legit_docx(src)

    result = _run(str(src))

    assert result.returncode == 0, (
        f"legit docx unexpectedly rejected\nstdout={result.stdout!r}\nstderr={result.stderr!r}"
    )
    payload = json.loads(result.stdout)
    assert payload["ok"] is True
    assert payload["format"] == "docx"
    assert "legitimate" in payload["text"]


def test_legitimate_xlsx_still_passes(tmp_path: Path) -> None:
    """Parallel control for the XLSX path — makes sure the zip-bomb
    guard doesn't false-positive on ordinary spreadsheets.
    """
    src = tmp_path / "legit.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "header"
    ws["B1"] = "value"
    wb.save(str(src))

    result = _run(str(src))

    assert result.returncode == 0, (
        f"legit xlsx unexpectedly rejected\nstdout={result.stdout!r}\nstderr={result.stderr!r}"
    )
    payload = json.loads(result.stdout)
    assert payload["ok"] is True
    assert payload["format"] == "xlsx"
    assert "header" in payload["text"]
